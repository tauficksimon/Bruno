#!/usr/bin/env python3
"""Build the reviewed Instantly launch files from the five Apollo exports.

The source workbooks contain an Apollo export label on row 1 and the real
headers on row 2, so they cannot be imported into Instantly as-is. This script
applies the campaign-v3 filters, removes explicitly catch-all addresses, keeps
one lead per company, and adds the persona metadata Bruno needs.
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "new-campaign"
OUTPUT_DIR = SOURCE_DIR / "instantly-imports"
BATCH = "2026-07-21"

US_VALUES = {"united states", "united states of america", "usa", "us", "u.s.", "u.s.a."}
TARGET_STATES = {
    "texas",
    "florida",
    "new york",
    "illinois",
    "georgia",
    "north carolina",
    "ohio",
    "pennsylvania",
    "new jersey",
    "tennessee",
    "tx",
    "fl",
    "ny",
    "il",
    "ga",
    "nc",
    "oh",
    "pa",
    "nj",
    "tn",
}


@dataclass(frozen=True)
class PersonaConfig:
    key: str
    source_glob: str
    output_file: str
    campaign_name: str
    persona: str
    target_role: str
    work_item: str
    employee_min: int
    employee_max: int
    industries: frozenset[str]
    include_titles: tuple[str, ...]
    exclude_titles: tuple[str, ...]
    geography: str


PERSONAS = (
    PersonaConfig(
        key="ea",
        source_glob="Persona 1 - EA*.xlsx",
        output_file="p1-ea.csv",
        campaign_name="Kinta | P1 EA | B1 | 2026-07",
        persona="EA",
        target_role="Executive Assistant",
        work_item="inbox and calendar",
        employee_min=21,
        employee_max=50,
        industries=frozenset({"real estate", "management consulting", "financial services", "insurance"}),
        include_titles=(
            r"\bfounder\b",
            r"\bco[- ]?founder\b",
            r"\bceo\b",
            r"chief executive officer",
            r"\bowner\b",
            r"\bcoo\b",
            r"chief operating officer",
        ),
        exclude_titles=(r"assistant", r"coordinator", r"office manager", r"\bintern\b"),
        geography="company_states",
    ),
    PersonaConfig(
        key="legal",
        source_glob="Persona 2 - Paralegal*.xlsx",
        output_file="p2-legal.csv",
        campaign_name="Kinta | P2 Legal | B1 | 2026-07",
        persona="Legal",
        target_role="Paralegal",
        work_item="case files",
        employee_min=2,
        employee_max=50,
        industries=frozenset({"law practice", "legal services"}),
        include_titles=(
            r"managing partner",
            r"founding partner",
            r"\bpartner\b",
            r"\bfounder\b",
            r"\bowner\b",
            r"principal attorney",
        ),
        exclude_titles=(
            r"paralegal",
            r"legal assistant",
            r"\bassociate\b",
            r"of counsel",
            r"\bclerk\b",
            r"assistant",
        ),
        geography="company_states",
    ),
    PersonaConfig(
        key="developer",
        source_glob="Persona 3 - Developers*.xlsx",
        output_file="p3-developer.csv",
        campaign_name="Kinta | P3 Developer | B1 | 2026-07",
        persona="Developer",
        target_role="Software Developer",
        work_item="repo",
        employee_min=11,
        employee_max=50,
        industries=frozenset({"computer software", "information technology & services", "internet"}),
        include_titles=(
            r"\bcto\b",
            r"chief technology officer",
            r"vp(?: of)? engineering",
            r"vp,\s*engineering",
            r"vice president(?: of)? engineering",
            r"head of engineering",
        ),
        exclude_titles=(r"assistant", r"recruiter", r"\btalent\b", r"\bhr\b", r"consultant"),
        geography="company_us",
    ),
    PersonaConfig(
        key="aec",
        source_glob="Persona 4 - Interior Design*.xlsx",
        output_file="p4-aec.csv",
        campaign_name="Kinta | P4 AEC | B1 | 2026-07",
        persona="AEC",
        target_role="BIM Modeler",
        work_item="project files",
        employee_min=2,
        employee_max=50,
        industries=frozenset({"architecture & planning", "civil engineering", "design"}),
        include_titles=(
            r"\bprincipal\b",
            r"principal architect",
            r"\bfounder\b",
            r"\bowner\b",
            r"managing partner",
            r"\bpresident\b",
            r"design director",
        ),
        exclude_titles=(
            r"assistant",
            r"drafter",
            r"\bjunior\b",
            r"\bintern\b",
            r"coordinator",
            r"project architect",
            r"bim modeler",
        ),
        geography="contact_us",
    ),
    PersonaConfig(
        key="social",
        source_glob="Persona 5 - Social Media*.xlsx",
        output_file="p5-social.csv",
        campaign_name="Kinta | P5 Social | B1 | 2026-07",
        persona="Marketing",
        target_role="Social Media Manager",
        work_item="content calendar",
        employee_min=5,
        employee_max=50,
        industries=frozenset(
            {"consumer goods", "retail", "apparel & fashion", "food & beverages", "marketing & advertising"}
        ),
        include_titles=(
            r"\bfounder\b",
            r"\bco[- ]?founder\b",
            r"\bceo\b",
            r"chief executive officer",
            r"\bowner\b",
            r"managing director",
            r"\bcmo\b",
            r"chief marketing officer",
            r"vp(?: of)? marketing",
            r"vice president(?: of)? marketing",
            r"head of marketing",
        ),
        exclude_titles=(
            r"social media manager",
            r"\bcontent\b",
            r"coordinator",
            r"specialist",
            r"\bintern\b",
            r"assistant",
        ),
        geography="contact_states",
    ),
)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value)).casefold()


def employee_count(value: Any) -> float | None:
    if value is None or text(value) == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        match = re.search(r"[\d,]+", text(value))
        return float(match.group(0).replace(",", "")) if match else None


def locate_source(config: PersonaConfig) -> Path:
    matches = sorted(SOURCE_DIR.glob(config.source_glob))
    if len(matches) != 1:
        raise RuntimeError(f"Expected one source for {config.key}, found {len(matches)}: {matches}")
    return matches[0]


def read_apollo_export(path: Path) -> list[dict[str, Any]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    headers = [text(value) for value in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    required = {
        "First Name",
        "Last Name",
        "Title",
        "Company Name",
        "Email",
        "Email Status",
        "Primary Email Catch-all Status",
        "# Employees",
        "Industry",
        "Country",
        "State",
        "Company Country",
        "Company State",
        "Qualify Contact",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"{path.name} is missing required columns: {', '.join(missing)}")

    return [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(min_row=3, values_only=True)
        if any(value is not None for value in row)
    ]


def geography_matches(config: PersonaConfig, row: dict[str, Any]) -> bool:
    contact_country = normalized(row.get("Country"))
    contact_state = normalized(row.get("State"))
    company_country = normalized(row.get("Company Country"))
    company_state = normalized(row.get("Company State"))

    if config.geography == "company_states":
        return company_country in {"", *US_VALUES} and company_state in {"", *TARGET_STATES}
    if config.geography == "company_us":
        return company_country in {"", *US_VALUES}
    if config.geography == "contact_us":
        return contact_country in US_VALUES
    if config.geography == "contact_states":
        return contact_country in US_VALUES and contact_state in TARGET_STATES
    raise RuntimeError(f"Unknown geography rule: {config.geography}")


def rejection_reasons(config: PersonaConfig, row: dict[str, Any]) -> list[str]:
    reasons: list[str] = []
    title = normalized(row.get("Title"))
    count = employee_count(row.get("# Employees"))

    if not text(row.get("Email")):
        reasons.append("missing_email")
    if normalized(row.get("Email Status")) != "verified":
        reasons.append("email_not_verified")
    if normalized(row.get("Qualify Contact")) == "disqualified":
        reasons.append("apollo_disqualified")
    if not any(re.search(pattern, title) for pattern in config.include_titles):
        reasons.append("title_not_included")
    if any(re.search(pattern, title) for pattern in config.exclude_titles):
        reasons.append("title_excluded")
    if not geography_matches(config, row):
        reasons.append("geography_mismatch")
    if count is None or not config.employee_min <= count <= config.employee_max:
        reasons.append("employee_count_mismatch")
    if normalized(row.get("Industry")) not in config.industries:
        reasons.append("industry_mismatch")
    if normalized(row.get("Primary Email Catch-all Status")) == "catch-all":
        reasons.append("catch_all")
    return reasons


def company_key(row: dict[str, Any]) -> str:
    email_domain = normalized(row.get("Email")).partition("@")[2]
    return normalized(row.get("Company Name")) or normalized(row.get("Website")) or email_domain


def csv_row(config: PersonaConfig, row: dict[str, Any]) -> dict[str, str]:
    return {
        "Email": text(row.get("Email")),
        "First Name": text(row.get("First Name")),
        "Last Name": text(row.get("Last Name")),
        "Company Name": text(row.get("Company Name")),
        "Title": text(row.get("Title")),
        "Website": text(row.get("Website")),
        "Persona": config.persona,
        "TargetRole": config.target_role,
        "WorkItem": config.work_item,
        "Batch": BATCH,
    }


def api_lead(config: PersonaConfig, row: dict[str, Any]) -> dict[str, Any]:
    return {
        "email": text(row.get("Email")),
        "first_name": text(row.get("First Name")),
        "last_name": text(row.get("Last Name")),
        "company_name": text(row.get("Company Name")),
        "job_title": text(row.get("Title")),
        "website": text(row.get("Website")),
        "phone": text(row.get("Mobile Phone")) or text(row.get("Work Direct Phone")),
        "custom_variables": {
            "persona": config.persona,
            "targetRole": config.target_role,
            "workItem": config.work_item,
            "batch": BATCH,
        },
    }


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest_personas: list[dict[str, Any]] = []
    audit_personas: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, str]] = []
    launch_emails: set[str] = set()

    csv_headers = [
        "Email",
        "First Name",
        "Last Name",
        "Company Name",
        "Title",
        "Website",
        "Persona",
        "TargetRole",
        "WorkItem",
        "Batch",
    ]

    for config in PERSONAS:
        source = locate_source(config)
        source_rows = read_apollo_export(source)
        accepted: list[dict[str, Any]] = []
        seen_companies: set[str] = set()
        reason_counts: dict[str, int] = {}
        catch_status_counts: dict[str, int] = {}

        for row in source_rows:
            reasons = rejection_reasons(config, row)
            key = company_key(row)
            if not reasons and key in seen_companies:
                reasons.append("duplicate_company")

            if reasons:
                for reason in set(reasons):
                    reason_counts[reason] = reason_counts.get(reason, 0) + 1
                rejected_rows.append(
                    {
                        "Persona": config.persona,
                        "Email": text(row.get("Email")),
                        "First Name": text(row.get("First Name")),
                        "Last Name": text(row.get("Last Name")),
                        "Company Name": text(row.get("Company Name")),
                        "Title": text(row.get("Title")),
                        "Reasons": ";".join(reasons),
                        "CatchAllStatus": text(row.get("Primary Email Catch-all Status")),
                        "CompanyCountry": text(row.get("Company Country")),
                        "CompanyState": text(row.get("Company State")),
                        "EmployeeCount": text(row.get("# Employees")),
                        "Industry": text(row.get("Industry")),
                    }
                )
                continue

            seen_companies.add(key)
            email = normalized(row.get("Email"))
            if email in launch_emails:
                raise RuntimeError(f"Cross-persona duplicate reached the launch pool: {email}")
            launch_emails.add(email)
            accepted.append(row)
            catch_status = normalized(row.get("Primary Email Catch-all Status")) or "unknown"
            catch_status_counts[catch_status] = catch_status_counts.get(catch_status, 0) + 1

        prepared_rows = [csv_row(config, row) for row in accepted]
        write_csv(OUTPUT_DIR / config.output_file, prepared_rows, csv_headers)
        manifest_personas.append(
            {
                "key": config.key,
                "campaignName": config.campaign_name,
                "persona": config.persona,
                "targetRole": config.target_role,
                "workItem": config.work_item,
                "sourceFile": source.name,
                "csvFile": config.output_file,
                "leadCount": len(accepted),
                "leads": [api_lead(config, row) for row in accepted],
            }
        )
        audit_personas.append(
            {
                "key": config.key,
                "persona": config.persona,
                "sourceFile": source.name,
                "sourceRows": len(source_rows),
                "launchLeads": len(accepted),
                "catchAllStatusInLaunchPool": dict(sorted(catch_status_counts.items())),
                "rejectionReasonCounts": dict(sorted(reason_counts.items())),
            }
        )

    rejected_headers = [
        "Persona",
        "Email",
        "First Name",
        "Last Name",
        "Company Name",
        "Title",
        "Reasons",
        "CatchAllStatus",
        "CompanyCountry",
        "CompanyState",
        "EmployeeCount",
        "Industry",
    ]
    write_csv(OUTPUT_DIR / "rejected-for-review.csv", rejected_rows, rejected_headers)

    generated_at = datetime.now(timezone.utc).isoformat()
    manifest = {
        "generatedAt": generated_at,
        "batch": BATCH,
        "totalLeadCount": len(launch_emails),
        "personas": manifest_personas,
    }
    audit = {
        "generatedAt": generated_at,
        "batch": BATCH,
        "sourceRowCount": sum(item["sourceRows"] for item in audit_personas),
        "launchLeadCount": len(launch_emails),
        "rejectedRowCount": len(rejected_rows),
        "personas": audit_personas,
        "notes": [
            "Explicitly catch-all addresses are held out of the launch pool.",
            "A blank Apollo catch-all status is retained as unknown; Instantly bounce protection remains enabled.",
            "Only one launch lead per normalized company name is retained.",
            "Campaigns remain paused after deployment.",
        ],
    }
    (OUTPUT_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    (OUTPUT_DIR / "audit.json").write_text(json.dumps(audit, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({"outputDir": str(OUTPUT_DIR), **audit}, indent=2))


if __name__ == "__main__":
    main()
